# services/user_service.py
import asyncio
import io
import math
import re
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta, timezone
from db.client import db
from prisma.enums import Role
from utils.exceptions import AppException
from utils.email import send_invitation_email, send_admin_invitation_email, send_promotion_email, send_demotion_email
from constants import ALLOWED_EMAIL_DOMAIN, DEFAULT_PAGE_SIZE, MAX_PAGE_SIZE

BULK_INVITE_WORKSHEET_NAME = "Invitations"


def _worksheet_for_bulk_invite(workbook: openpyxl.Workbook):
    if BULK_INVITE_WORKSHEET_NAME in workbook.sheetnames:
        return workbook[BULK_INVITE_WORKSHEET_NAME]
    return workbook.active


def build_bulk_invite_template_workbook_bytes() -> bytes:
    """
    .xlsx with a data sheet matching bulk_invite_users. Instructions on a second sheet.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = BULK_INVITE_WORKSHEET_NAME
    headers = ["email", "name", "designation", "dateOfJoining", "birthdate"]
    ws.append(headers)
    ws.freeze_panes = "A2"
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 22

    help_ws = wb.create_sheet("Instructions", 1)
    help_ws["A1"] = "Bulk invite"
    help_ws["A2"] = (
        f"1) Enter invitees on the '{BULK_INVITE_WORKSHEET_NAME}' sheet starting in row 2. "
        "2) Column 'email' is required. "
        f"3) Only addresses ending with @{ALLOWED_EMAIL_DOMAIN} are accepted. "
        "4) Optional columns: name, designation, dateOfJoining, birthdate. "
        "The legacy header 'birthday' is still accepted if you reuse an old file. "
        "5) Use Excel date cells or ISO dates (e.g. 2025-01-15). "
        "6) Save as .xlsx and upload."
    )
    help_ws.column_dimensions["A"].width = 100
    help_ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _is_user_banned_for_display(user) -> bool:
    """
    True when the user has an active time-bound ban for display purposes.
    Elapsed bans and isBanned without bannedUntil are not shown as banned.
    """
    if not user.isBanned or user.bannedUntil is None:
        return False
    now = datetime.now(timezone.utc)
    banned_until = (
        user.bannedUntil.replace(tzinfo=timezone.utc)
        if user.bannedUntil.tzinfo is None
        else user.bannedUntil
    )
    return now < banned_until


def derive_app_status(user) -> str:
    """App surface: deleted, banned, invited until first app OAuth, else active."""
    if user.deletedAt is not None:
        return "DELETED"
    if _is_user_banned_for_display(user):
        return "BANNED"
    if not user.hasLoggedInApp:
        return "INVITED"
    return "ACTIVE"


def derive_panel_status(user) -> str:
    """Panel surface: deleted, banned, invited until first panel OAuth, else active."""
    if user.deletedAt is not None:
        return "DELETED"
    if _is_user_banned_for_display(user):
        return "BANNED"
    if not user.hasLoggedInPanel:
        return "INVITED"
    return "ACTIVE"


def serialize_user(u) -> dict:
    return {
        "id": u.id,
        "email": u.email,
        "name": u.name,
        "picture": u.picture,
        "designation": u.designation,
        "dateOfJoining": u.dateOfJoining.isoformat() if u.dateOfJoining else None,
        "birthdate": u.birthdate.isoformat() if u.birthdate else None,
        "role": str(u.role),
        "appStatus": derive_app_status(u),
        "panelStatus": derive_panel_status(u),
        "isBanned": u.isBanned,
        "bannedUntil": u.bannedUntil.isoformat() if u.bannedUntil else None,
        "banReason": u.banReason,
        "hasLoggedInApp": u.hasLoggedInApp,
        "hasLoggedInPanel": u.hasLoggedInPanel,
        "createdAt": u.createdAt.isoformat(),
        "deletedAt": u.deletedAt.isoformat() if u.deletedAt else None,
    }


# ============================================================
# INVITE
# ============================================================

async def invite_users(people: list) -> dict:
    """Invite users to the app. Creates MEMBER stub records with optional profile data."""
    return await _process_person_invites(people, role=Role.MEMBER)


async def invite_admins(people: list) -> dict:
    """Invite admins to the panel. Creates ADMIN stub records. SUPER_ADMIN only."""
    return await _process_person_invites(people, role=Role.ADMIN)


async def _batch_invite_create(
    create_payloads: list[dict],
    skipped: list[dict],
) -> tuple[list[dict], list[str]]:
    """
    Prefetch existing emails, create_many new users, return invited rows and emails to send.
    """
    invited: list[dict] = []
    emails_to_send: list[str] = []

    if not create_payloads:
        return invited, emails_to_send

    all_emails = [p["email"] for p in create_payloads]
    existing_users = await db.user.find_many(where={"email": {"in": all_emails}})
    existing_emails = {u.email for u in existing_users}

    to_create: list[dict] = []
    for create_data in create_payloads:
        email = create_data["email"]
        if email in existing_emails:
            skipped.append({"email": email, "reason": "User already exists"})
            continue
        to_create.append(create_data)

    if to_create:
        await db.user.create_many(data=to_create)
        for create_data in to_create:
            invited.append({"email": create_data["email"]})
            emails_to_send.append(create_data["email"])

    return invited, emails_to_send


async def _process_person_invites(people: list, role: Role) -> dict:
    """
    Process invite for a list of person objects.
    Each person has email (required) + optional name, designation, dateOfJoining, birthdate.
    """
    invited: list[dict] = []
    skipped: list[dict] = []
    create_payloads: list[dict] = []

    for person in people:
        email = str(person.email)

        if not email.endswith(f"@{ALLOWED_EMAIL_DOMAIN}"):
            skipped.append({
                "email": email,
                "reason": f"Only @{ALLOWED_EMAIL_DOMAIN} emails can be invited"
            })
            continue

        create_data: dict = {"email": email, "role": role}

        if person.name and str(person.name).strip():
            create_data["name"] = str(person.name).strip()
        if person.designation and str(person.designation).strip():
            create_data["designation"] = str(person.designation).strip()
        if person.dateOfJoining:
            create_data["dateOfJoining"] = person.dateOfJoining
        if person.birthdate:
            create_data["birthdate"] = person.birthdate

        create_payloads.append(create_data)

    batch_invited, emails_to_send = await _batch_invite_create(create_payloads, skipped)
    invited.extend(batch_invited)

    return {
        "invited": invited,
        "skipped": skipped,
        "emailsToSend": emails_to_send,
    }


def _parse_excel_rows(file_bytes: bytes) -> tuple[list[dict], list[dict]]:
    """
    CPU-bound Excel parse — run via asyncio.to_thread from async handlers.
    Returns (valid_row_dicts, skipped_entries).
    """
    try:
        workbook = openpyxl.load_workbook(io.BytesIO(file_bytes))
    except Exception:
        raise AppException(400, "Invalid Excel file")

    sheet = _worksheet_for_bulk_invite(workbook)

    raw_headers = [cell.value for cell in sheet[1]]
    header_map = {}
    for i, header in enumerate(raw_headers):
        if header:
            header_map[str(header).strip().lower()] = i

    if "email" not in header_map:
        raise AppException(400, "Email column not found in the Excel file")

    email_idx = header_map.get("email", -1)
    name_idx = header_map.get("name", -1)
    designation_idx = header_map.get("designation", -1)
    doj_idx = header_map.get("dateofjoining", -1)
    birthdate_idx = header_map.get("birthdate", header_map.get("birthday", -1))

    def get_cell(row, idx):
        if idx == -1 or idx >= len(row):
            return None
        val = row[idx]
        if val is None:
            return None
        return str(val).strip() if not isinstance(val, datetime) else val

    def parse_date(val):
        if val is None:
            return None
        if isinstance(val, datetime):
            return val
        try:
            return datetime.fromisoformat(str(val))
        except Exception:
            return None

    email_regex = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")
    raw_rows: list[dict] = []
    skipped: list[dict] = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        email_val = get_cell(row, email_idx)
        if not email_val:
            continue

        if not email_regex.match(email_val):
            skipped.append({"email": email_val, "reason": "Invalid email format"})
            continue

        raw_rows.append({
            "email": email_val,
            "name": get_cell(row, name_idx),
            "designation": get_cell(row, designation_idx),
            "dateOfJoining": parse_date(get_cell(row, doj_idx)),
            "birthdate": parse_date(get_cell(row, birthdate_idx)),
        })

    return raw_rows, skipped


async def bulk_invite_users(file_bytes: bytes, filename: str, role: Role = Role.MEMBER) -> dict:
    """
    Bulk invite from Excel file.
    Expected columns: email (required), name, designation, dateOfJoining, birthdate (all optional).
    role param determines whether inviting users (MEMBER) or admins (ADMIN).
    """
    if not filename.lower().endswith(".xlsx") and not filename.lower().endswith(".xls"):
        raise AppException(400, "Invalid file format. Only .xlsx and .xls files are allowed")

    raw_rows, skipped = await asyncio.to_thread(_parse_excel_rows, file_bytes)

    if not raw_rows:
        raise AppException(400, "No valid emails found in the Excel file")

    create_payloads: list[dict] = []
    for row_data in raw_rows:
        email = row_data["email"]

        if not email.endswith(f"@{ALLOWED_EMAIL_DOMAIN}"):
            skipped.append({
                "email": email,
                "reason": f"Only @{ALLOWED_EMAIL_DOMAIN} emails can be invited",
            })
            continue

        create_data: dict = {"email": email, "role": role}

        if row_data["name"]:
            create_data["name"] = row_data["name"]
        if row_data["designation"]:
            create_data["designation"] = row_data["designation"]
        if row_data["dateOfJoining"]:
            create_data["dateOfJoining"] = row_data["dateOfJoining"]
        if row_data["birthdate"]:
            create_data["birthdate"] = row_data["birthdate"]

        create_payloads.append(create_data)

    invited, emails_to_send = await _batch_invite_create(create_payloads, skipped)

    return {
        "invited": invited,
        "skipped": skipped,
        "emailsToSend": emails_to_send,
        "totalProcessed": len(raw_rows) + len([s for s in skipped if "Invalid email format" in s.get("reason", "")]),
        "totalInvited": len(invited),
        "totalSkipped": len(skipped),
    }


async def resend_invite(emails: list) -> dict:
    """
    Resend invite emails to existing users/admins.
    No new DB entry created.
    Sends admin email template if role is ADMIN, user template otherwise.
    """
    sent = []
    skipped = []

    for email in emails:
        user = await db.user.find_unique(where={"email": email})

        if not user:
            skipped.append({"email": email, "reason": "User not found"})
            continue

        if user.role == Role.SUPER_ADMIN:
            raise AppException(
                400,
                "Super admins are hardcoded in the system and cannot be re-invited.",
            )

        if user.role == Role.ADMIN:
            await send_admin_invitation_email(email)
        else:
            await send_invitation_email(email)

        sent.append({"email": email})

    return {
        "sent": sent,
        "skipped": skipped
    }


# ============================================================
# DELETE
# ============================================================

async def _check_delete_permission(caller, target_user):
    # Block self deletion
    if caller.id == target_user.id:
        raise AppException(400, "You cannot delete yourself")

    # Nobody can delete a SUPER_ADMIN
    if target_user.role == Role.SUPER_ADMIN:
        raise AppException(403, "Super Admins cannot be deleted")

    # ADMIN can only delete MEMBERs — only SUPER_ADMIN can delete ADMINs
    if caller.role == Role.ADMIN and target_user.role != Role.MEMBER:
        raise AppException(403, "Admins can only delete members")


async def delete_user(current_user, target_user_id: str) -> dict:
    target_user = await db.user.find_unique(where={"id": target_user_id})
    if not target_user:
        raise AppException(404, "User not found")
    if target_user.deletedAt is not None:
        raise AppException(400, "User is already deleted")
    await _check_delete_permission(current_user, target_user)
    now = datetime.now(timezone.utc)
    await db.user.update(
        where={"id": target_user_id},
        data={"deletedAt": now}
    )
    return {"deletedUserId": target_user_id}


async def bulk_delete_users(current_user, user_ids: list) -> dict:
    deleted = []
    skipped = []

    for user_id in user_ids:
        target_user = await db.user.find_unique(where={"id": user_id})

        if not target_user:
            skipped.append({"userId": user_id, "reason": "User not found"})
            continue

        if target_user.deletedAt is not None:
            skipped.append({"userId": user_id, "reason": "User already deleted"})
            continue

        try:
            await _check_delete_permission(current_user, target_user)
        except AppException as e:
            skipped.append({"userId": user_id, "reason": e.message})
            continue

        now = datetime.now(timezone.utc)
        await db.user.update(
            where={"id": user_id},
            data={"deletedAt": now}
        )
        deleted.append({"userId": user_id})

    return {"deleted": deleted, "skipped": skipped}


async def get_deleted_users() -> list:
    users = await db.user.find_many(
        where={"deletedAt": {"not": None}},
        order={"deletedAt": "desc"}
    )
    return [serialize_user(u) for u in users]


async def restore_user(current_user, target_user_id: str) -> dict:
    target_user = await db.user.find_unique(where={"id": target_user_id})
    if not target_user:
        raise AppException(404, "User not found")
    if target_user.deletedAt is None:
        raise AppException(400, "User is not deleted")
    if target_user.role == Role.SUPER_ADMIN:
        raise AppException(403, "Super Admin cannot be restored through this endpoint")
    if current_user.role == Role.ADMIN and target_user.role != Role.MEMBER:
        raise AppException(403, "Admins can only restore members")
    restored_user = await db.user.update(
        where={"id": target_user_id},
        data={"deletedAt": None}
    )
    return serialize_user(restored_user)


# ============================================================
# BAN
# ============================================================

def _check_ban_permission(caller, target_user):
    """Shared permission check for ban, edit ban, unban."""
    if caller.id == target_user.id:
        raise AppException(400, "You cannot perform this action on yourself")

    if target_user.role == Role.SUPER_ADMIN:
        raise AppException(403, "Super Admins cannot be banned")

    # ADMIN cannot perform any ban operation on another ADMIN
    if caller.role == Role.ADMIN and target_user.role == Role.ADMIN:
        raise AppException(403, "Admins cannot ban other admins")


async def ban_user(current_user, target_user_id: str, duration_hours: int, reason: str) -> dict:
    target_user = await db.user.find_unique(where={"id": target_user_id})

    if not target_user:
        raise AppException(404, "User not found")

    _check_ban_permission(current_user, target_user)

    if duration_hours <= 0:
        raise AppException(400, "Ban duration must be greater than 0 hours")

    banned_until = datetime.now(timezone.utc) + timedelta(hours=duration_hours)

    updated_user = await db.user.update(
        where={"id": target_user_id},
        data={
            "isBanned": True,
            "bannedUntil": banned_until,
            "banReason": reason,
        }
    )

    return {
        "userId": updated_user.id,
        "isBanned": updated_user.isBanned,
        "bannedUntil": updated_user.bannedUntil.isoformat(),
        "banReason": updated_user.banReason,
    }


async def edit_ban(current_user, target_user_id: str, duration_hours: int, ban_updates: dict | None = None) -> dict:
    if current_user.id == target_user_id:
        raise AppException(400, "You cannot edit your own ban")

    target_user = await db.user.find_unique(where={"id": target_user_id})

    if not target_user:
        raise AppException(404, "User not found")

    _check_ban_permission(current_user, target_user)

    if not target_user.isBanned:
        raise AppException(400, "This user is not currently banned")

    if duration_hours <= 0:
        raise AppException(400, "Ban duration must be greater than 0 hours")

    banned_until = datetime.now(timezone.utc) + timedelta(hours=duration_hours)

    update_data: dict = {"bannedUntil": banned_until}
    if ban_updates and "reason" in ban_updates:
        update_data["banReason"] = ban_updates["reason"]

    updated_user = await db.user.update(
        where={"id": target_user_id},
        data=update_data
    )

    return {
        "userId": updated_user.id,
        "isBanned": updated_user.isBanned,
        "bannedUntil": updated_user.bannedUntil.isoformat(),
        "banReason": updated_user.banReason,
    }


async def unban_user(current_user, target_user_id: str) -> dict:
    if current_user.id == target_user_id:
        raise AppException(400, "You cannot unban yourself")

    target_user = await db.user.find_unique(where={"id": target_user_id})

    if not target_user:
        raise AppException(404, "User not found")

    _check_ban_permission(current_user, target_user)

    if not target_user.isBanned:
        raise AppException(400, "This user is not currently banned")

    await db.user.update(
        where={"id": target_user_id},
        data={
            "isBanned": False,
            "bannedUntil": None,
            "banReason": None,
        }
    )

    return {"userId": target_user_id, "isBanned": False}


# ============================================================
# ROLE CHANGE
# ============================================================

async def change_role(current_user, target_user_id: str, new_role: str) -> dict:
    """
    Change an admin's role to MEMBER or a member's role to ADMIN.
    SUPER_ADMIN only. Cannot change SUPER_ADMIN role.
    Accepted values for new_role: 'MEMBER', 'ADMIN'
    """
    if current_user.id == target_user_id:
        raise AppException(400, "You cannot change your own role")

    if new_role not in ("MEMBER", "ADMIN"):
        raise AppException(400, "Role must be either 'MEMBER' or 'ADMIN'")

    target_user = await db.user.find_unique(where={"id": target_user_id})

    if not target_user:
        raise AppException(404, "User not found")

    if target_user.role == Role.SUPER_ADMIN:
        raise AppException(403, "Super Admin role cannot be changed")

    if str(target_user.role) == new_role:
        raise AppException(400, f"User already has the role {new_role}")

    updated_user = await db.user.update(
        where={"id": target_user_id},
        data={"role": Role[new_role]}
    )

    if new_role == "ADMIN":
        await send_promotion_email(updated_user.email)
    elif new_role == "MEMBER":
        await send_demotion_email(updated_user.email)

    return {
        "userId": updated_user.id,
        "role": str(updated_user.role)
    }


# ============================================================
# EDIT USER
# ============================================================
async def edit_user(current_user, target_user_id: str, updates: dict) -> dict:
    """
    Edit profile fields of a user.
    ADMIN can only edit MEMBERs.
    SUPER_ADMIN can edit MEMBERs and ADMINs.
    Nobody can edit a SUPER_ADMIN.
    Only provided fields are updated (partial update).
    """
    if current_user.id == target_user_id:
        raise AppException(400, "You cannot edit your own profile through this endpoint")

    target_user = await db.user.find_unique(where={"id": target_user_id})
    if not target_user:
        raise AppException(404, "User not found")

    if target_user.role == Role.SUPER_ADMIN:
        raise AppException(403, "Super Admin profile cannot be edited")

    if current_user.role == Role.ADMIN and target_user.role != Role.MEMBER:
        raise AppException(403, "Admins can only edit members")

    # Build update payload — only include keys that were explicitly provided
    update_data = {k: v for k, v in updates.items() if v is not None}

    if not update_data:
        raise AppException(400, "No valid fields provided for update")

    updated_user = await db.user.update(
        where={"id": target_user_id},
        data=update_data
    )

    return serialize_user(updated_user)


# ============================================================
# GET ALL USERS
# ============================================================

VALID_LIST_STATUSES = frozenset({"ACTIVE", "INVITED", "BANNED", "DELETED"})


VALID_STATUS_SURFACES = frozenset({"app", "panel"})


async def get_all_users(
    search: str = None,
    status: str = None,
    status_surface: str = "app",
    role: str = None,
    page: int = 1,
    page_size: int = DEFAULT_PAGE_SIZE,
) -> dict:
    if status is not None and status not in VALID_LIST_STATUSES:
        raise AppException(
            400,
            f"Invalid status. Must be one of: {', '.join(sorted(VALID_LIST_STATUSES))}",
        )
    if status_surface not in VALID_STATUS_SURFACES:
        raise AppException(400, "Invalid status_surface. Must be 'app' or 'panel'.")

    page_size = min(page_size, MAX_PAGE_SIZE)
    skip = (page - 1) * page_size

    # Always start with deletedAt filter
    where: dict = {}

    if status == "DELETED":
        where["deletedAt"] = {"not": None}
    else:
        where["deletedAt"] = None

        # Translate derived status into real DB column conditions
        if status == "ACTIVE":
            login_field = "hasLoggedInPanel" if status_surface == "panel" else "hasLoggedInApp"
            where[login_field] = True
            where["isBanned"] = False

        elif status == "INVITED":
            login_field = "hasLoggedInPanel" if status_surface == "panel" else "hasLoggedInApp"
            where[login_field] = False
            where["isBanned"] = False

        elif status == "BANNED":
            where["isBanned"] = True
            where["bannedUntil"] = {
                "gt": datetime.now(timezone.utc)
            }

    # Role filter
    if role is not None:
        where["role"] = role

    # Search filter
    if search:
        where["OR"] = [
            {"email": {"contains": search, "mode": "insensitive"}},
            {"name": {"contains": search, "mode": "insensitive"}},
        ]

    # Run count and find_many in parallel — single round trip
    total, users = await asyncio.gather(
        db.user.count(where=where),
        db.user.find_many(
            where=where,
            skip=skip,
            take=page_size,
            order={"createdAt": "desc"},
        ),
    )

    total_pages = math.ceil(total / page_size) if total else 0

    return {
        "users": [serialize_user(u) for u in users],
        "total": total,
        "page": page,
        "pageSize": page_size,
        "totalPages": total_pages,
    }